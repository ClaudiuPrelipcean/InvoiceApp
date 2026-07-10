import os
import re
import shutil
import tempfile
import zipfile
from io import BytesIO

import pandas as pd
import pytesseract
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from pdf2image import convert_from_path


# ============================================================
# CONFIGURARE PAGINĂ
# ============================================================

st.set_page_config(
    page_title="Procesator Facturi",
    page_icon="📄",
    layout="wide"
)

st.title("📄 Procesator automat de facturi")
st.write(
    "Încarcă facturile PDF, iar aplicația le redenumește, "
    "le sortează și generează automat fișierul Excel."
)


# ============================================================
# SETĂRI
# ============================================================

st.sidebar.header("⚙️ Setări aplicație")

firma_proprie = st.sidebar.text_input(
    "Firma proprie",
    value="Company Name"
)

tesseract_path = st.sidebar.text_input(
    "Calea către Tesseract",
    value=r"C:\Program Files\Tesseract-OCR\tesseract.exe"
)

poppler_path = st.sidebar.text_input(
    "Calea către folderul bin din Poppler",
    value=r"C:\poppler\Library\bin"
)

st.sidebar.info(
    "Aplicația rulează local și are nevoie de Tesseract OCR și Poppler instalate."
)


# ============================================================
# ZONELE DIN FACTURĂ FOLOSITE PENTRU OCR
# ============================================================

# Vânzător
VENDOR_BOX = (50, 50, 700, 400)

# Număr factură și dată
DATE_BOX = (700, 50, 1400, 650)

# Cumpărător
BUYER_BOX = (1400, 50, 2000, 400)

# Valorile nete
VALUE_BOX = (2120, 700, 2400, 1700)


# ============================================================
# FUNCȚII GENERALE
# ============================================================

def normalize_company_name(company_name):
    """
    Normalizează numele firmei pentru comparații.
    Elimină spațiile, punctele și alte caractere.
    """

    return re.sub(r"[^A-Z0-9]", "", company_name.upper())


def clean_filename(text):
    """
    Curăță textul pentru a putea fi folosit într-un nume de fișier.
    """

    text = str(text).strip()

    # Elimină caracterele interzise în numele fișierelor Windows
    text = re.sub(r'[<>:"/\\|?*]', "", text)

    # Păstrează comportamentul codului original
    text = text.replace(" ", "_")
    text = text.replace(".", "")
    text = text.replace("-", "_")

    # Elimină underscore-urile multiple
    text = re.sub(r"_+", "_", text)

    text = text.strip("_")

    if not text:
        return "Necunoscut"

    return text[:150]


def clean_sheet_name(text):
    """
    Curăță numele pentru a putea fi folosit ca nume de sheet Excel.
    
    """

    text = str(text)

    # Caractere interzise în numele sheet-urilor Excel
    text = re.sub(r"[\[\]:*?/\\]", "", text)

    text = text.strip()

    if not text:
        text = "Necunoscut"

    return text[:31]


def get_unique_file_path(folder, filename):
    """
    Evită suprascrierea fișierelor care au același nume.
    """

    destination_path = os.path.join(folder, filename)

    if not os.path.exists(destination_path):
        return destination_path

    base_name, extension = os.path.splitext(filename)
    counter = 2

    while True:
        new_filename = f"{base_name}_{counter}{extension}"
        destination_path = os.path.join(folder, new_filename)

        if not os.path.exists(destination_path):
            return destination_path

        counter += 1


# ============================================================
# FUNCȚII OCR
# ============================================================

def convert_pdf_first_page(pdf_path):
    """
    Convertește doar prima pagină a PDF-ului într-o imagine.
    """

    pages = convert_from_path(
        pdf_path,
        dpi=200,
        first_page=1,
        last_page=1,
        poppler_path=poppler_path
    )

    if not pages:
        raise ValueError("PDF-ul nu conține nicio pagină.")

    return pages[0]


def extract_text_from_page(page):
    """
    Extrage prin OCR:
    - vânzătorul;
    - numărul și data facturii;
    - cumpărătorul.
    """

    vendor_img = page.crop(VENDOR_BOX)
    date_img = page.crop(DATE_BOX)
    buyer_img = page.crop(BUYER_BOX)

    vendor_text = pytesseract.image_to_string(
        vendor_img,
        lang="ron+eng"
    )

    date_text = pytesseract.image_to_string(
        date_img,
        lang="ron+eng"
    )

    buyer_text = pytesseract.image_to_string(
        buyer_img,
        lang="ron+eng"
    )

    vendor_text = " ".join(vendor_text.split())
    date_text = " ".join(date_text.split())
    buyer_text = " ".join(buyer_text.split())

    return vendor_text, date_text, buyer_text


def extract_vendor(vendor_text):
    """
    Extrage numele vânzătorului.
    """

    patterns = [
        r"Nume\s+(.+?)\s+Nr\.?\s*inregistrare",
        r"Nume\s+(.+?)\s+Nr\.?\s*înregistrare",
        r"Nume\s+(.+?)\s+CUI",
    ]

    for pattern in patterns:
        match = re.search(
            pattern,
            vendor_text,
            re.IGNORECASE
        )

        if match:
            return match.group(1).strip()

    return "Vanzator_Necunoscut"


def extract_buyer(buyer_text):
    """
    Extrage numele cumpărătorului.
    """

    patterns = [
        r"Nume\s+(.+?)\s+Nr\.?\s*inregistrare",
        r"Nume\s+(.+?)\s+Nr\.?\s*înregistrare",
        r"Nume\s+(.+?)\s+CUI",
    ]

    for pattern in patterns:
        match = re.search(
            pattern,
            buyer_text,
            re.IGNORECASE
        )

        if match:
            return match.group(1).strip()

    return "Cumparator_Necunoscut"


def extract_invoice_number(date_text):
    """
    Extrage numărul facturii.
    """

    patterns = [
        r"Nr\.?\s*factura\s+(.+?)\s+Codul\s+tipului",
        r"Nr\.?\s*factură\s+(.+?)\s+Codul\s+tipului",
        r"Numar\s+factura\s+(.+?)\s+Codul\s+tipului",
    ]

    for pattern in patterns:
        match = re.search(
            pattern,
            date_text,
            re.IGNORECASE
        )

        if match:
            return match.group(1).strip()

    return "Nr_Necunoscut"


def extract_invoice_date(date_text):
    """
    Extrage data emiterii facturii.
    Acceptă formatele:
    - 2026-07-11
    - 11.07.2026
    - 11/07/2026
    """

    match = re.search(
        r"Data\s+emitere\s+([0-9]{4}-[0-9]{2}-[0-9]{2})",
        date_text,
        re.IGNORECASE
    )

    if match:
        return match.group(1).strip()

    match = re.search(
        r"Data\s+emitere\s+([0-9]{2}[./-][0-9]{2}[./-][0-9]{4})",
        date_text,
        re.IGNORECASE
    )

    if match:
        date_value = match.group(1).strip()
        date_value = date_value.replace("/", ".").replace("-", ".")

        try:
            day, month, year = date_value.split(".")
            return f"{year}-{month}-{day}"
        except ValueError:
            return date_value

    return "Data_Necunoscuta"


# ============================================================
# EXTRAGERE VALOARE NETĂ
# ============================================================

def parse_number(number_text):
    """
    Transformă diferite formate numerice în float.
    
    """

    number_text = number_text.strip()
    number_text = number_text.replace(" ", "")

    if "," in number_text and "." in number_text:
        if number_text.rfind(",") > number_text.rfind("."):
            # Format românesc: 1.234,56
            number_text = number_text.replace(".", "")
            number_text = number_text.replace(",", ".")
        else:
            # Format englezesc: 1,234.56
            number_text = number_text.replace(",", "")

    elif "," in number_text:
        # Format: 1234,56
        number_text = number_text.replace(",", ".")

    return float(number_text)


def extract_net_value(page):
    """
    Extrage și adună valorile nete din zona stabilită.
    """

    value_img = page.crop(VALUE_BOX)

    value_text = pytesseract.image_to_string(
        value_img,
        lang="ron+eng",
        config="--psm 6"
    )

    value_text = " ".join(value_text.split())

    # Caută numere de forma:
    # 123
    # 123,45
    # 1.234,56
    # 1,234.56
    number_matches = re.findall(
        r"(?<!\d)\d{1,3}(?:[.\s]\d{3})*(?:,\d{1,2})?"
        r"|(?<!\d)\d+(?:[.,]\d{1,2})?",
        value_text
    )

    values = []

    for number_text in number_matches:
        try:
            value = parse_number(number_text)

            # Ignorăm valori negative sau imposibile
            if value >= 0:
                values.append(value)

        except ValueError:
            continue

    total = round(sum(values), 2)

    return total, value_text


# ============================================================
# CREARE EXCEL
# ============================================================

def get_unique_sheet_name(workbook, desired_name):
    """
    Creează un nume unic pentru sheet.
    """

    sheet_name = clean_sheet_name(desired_name)

    if sheet_name not in workbook.sheetnames:
        return sheet_name

    counter = 2

    while True:
        suffix = f"_{counter}"
        shortened_name = sheet_name[:31 - len(suffix)]
        candidate = f"{shortened_name}{suffix}"

        if candidate not in workbook.sheetnames:
            return candidate

        counter += 1


def create_excel(results):
    """
    Creează fișierul Excel în memorie.
    Fiecare vânzător primește propriul sheet.
    """

    workbook = Workbook()

    # Elimină sheet-ul creat automat
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    vendor_sheet_map = {}

    for result in results:
        vendor = result["Vanzator"]

        if vendor not in vendor_sheet_map:
            sheet_name = get_unique_sheet_name(
                workbook,
                clean_filename(vendor)
            )

            vendor_sheet_map[vendor] = sheet_name

            worksheet = workbook.create_sheet(sheet_name)

            worksheet.append([
                "Data Factura",
                "Numar Factura",
                "Total Valoare Neta"
            ])

            # Formatarea antetului
            header_fill = PatternFill(
                fill_type="solid",
                fgColor="D9EAF7"
            )

            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.fill = header_fill

        else:
            sheet_name = vendor_sheet_map[vendor]
            worksheet = workbook[sheet_name]

        worksheet.append([
            result["Data factura"],
            result["Numar factura"],
            result["Valoare neta"]
        ])

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]

        last_data_row = worksheet.max_row
        total_row = last_data_row + 2

        worksheet[f"B{total_row}"] = "TOTAL FACTURI:"
        worksheet[f"B{total_row}"].font = Font(bold=True)

        worksheet[f"C{total_row}"] = (
            f"=SUM(C2:C{last_data_row})"
        )

        worksheet[f"C{total_row}"].font = Font(bold=True)
        worksheet[f"C{total_row}"].number_format = '#,##0.00'

        # Formatează valorile numerice
        for row_number in range(2, last_data_row + 1):
            worksheet[f"C{row_number}"].number_format = '#,##0.00'

        # Ajustează lățimea coloanelor
        for column_number, column_cells in enumerate(
            worksheet.columns,
            start=1
        ):
            maximum_length = 0

            for cell in column_cells:
                if cell.value is not None:
                    maximum_length = max(
                        maximum_length,
                        len(str(cell.value))
                    )

            worksheet.column_dimensions[
                get_column_letter(column_number)
            ].width = min(maximum_length + 3, 50)

    excel_buffer = BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)

    return excel_buffer.getvalue()


# ============================================================
# CREARE ARHIVĂ ZIP
# ============================================================

def create_zip(folder_path):
    """
    Creează în memorie arhiva ZIP cu facturile sortate.
    """

    zip_buffer = BytesIO()

    with zipfile.ZipFile(
        zip_buffer,
        mode="w",
        compression=zipfile.ZIP_DEFLATED
    ) as zip_file:

        for root, directories, files in os.walk(folder_path):
            for filename in files:
                full_path = os.path.join(root, filename)

                archive_path = os.path.relpath(
                    full_path,
                    folder_path
                )

                zip_file.write(
                    full_path,
                    archive_path
                )

    zip_buffer.seek(0)

    return zip_buffer.getvalue()


# ============================================================
# VERIFICAREA CONFIGURĂRII
# ============================================================

configuration_errors = []

if not os.path.isfile(tesseract_path):
    configuration_errors.append(
        "Nu a fost găsit fișierul Tesseract la calea introdusă."
    )

if not os.path.isdir(poppler_path):
    configuration_errors.append(
        "Nu a fost găsit folderul Poppler la calea introdusă."
    )

if configuration_errors:
    for error in configuration_errors:
        st.warning(error)


# ============================================================
# UPLOAD FACTURI
# ============================================================

uploaded_files = st.file_uploader(
    "Încarcă facturile PDF",
    type=["pdf"],
    accept_multiple_files=True,
    help="Poți selecta mai multe facturi PDF în același timp."
)

if uploaded_files:
    st.success(
        f"Au fost încărcate {len(uploaded_files)} facturi."
    )

    with st.expander("Vezi fișierele încărcate"):
        for uploaded_file in uploaded_files:
            st.write(f"• {uploaded_file.name}")


# ============================================================
# PROCESAREA FACTURILOR
# ============================================================

process_button = st.button(
    "▶️ Procesează facturile",
    type="primary",
    use_container_width=True
)

if process_button:

    if not uploaded_files:
        st.error("Încarcă cel puțin o factură PDF.")
        st.stop()

    if configuration_errors:
        st.error(
            "Corectează mai întâi căile către Tesseract și Poppler."
        )
        st.stop()

    pytesseract.pytesseract.tesseract_cmd = tesseract_path

    results = []
    errors = []
    logs = []

    with tempfile.TemporaryDirectory() as temporary_folder:

        raw_folder = os.path.join(
            temporary_folder,
            "FacturiRaw"
        )

        processed_folder = os.path.join(
            temporary_folder,
            "FacturiPrelucrate"
        )

        clienti_folder = os.path.join(
            processed_folder,
            "FacturiClienti"
        )

        furnizori_folder = os.path.join(
            processed_folder,
            "FacturiFurnizori"
        )

        os.makedirs(raw_folder, exist_ok=True)
        os.makedirs(clienti_folder, exist_ok=True)
        os.makedirs(furnizori_folder, exist_ok=True)

        progress_bar = st.progress(0)
        status_text = st.empty()

        number_of_files = len(uploaded_files)

        for index, uploaded_file in enumerate(
            uploaded_files,
            start=1
        ):
            status_text.write(
                f"Procesez: **{uploaded_file.name}**"
            )

            original_pdf_path = os.path.join(
                raw_folder,
                uploaded_file.name
            )

            with open(original_pdf_path, "wb") as output_file:
                output_file.write(uploaded_file.getbuffer())

            try:
                # Convertește PDF-ul o singură dată
                first_page = convert_pdf_first_page(
                    original_pdf_path
                )

                # OCR date factură
                vendor_text, date_text, buyer_text = (
                    extract_text_from_page(first_page)
                )

                vendor = extract_vendor(vendor_text)
                buyer = extract_buyer(buyer_text)
                invoice_number = extract_invoice_number(date_text)
                invoice_date = extract_invoice_date(date_text)

                # OCR valoare netă
                net_value, value_ocr_text = extract_net_value(
                    first_page
                )

                vendor_clean = clean_filename(vendor)
                invoice_number_clean = clean_filename(
                    invoice_number
                )

                # Logica originală:
                # numele fișierului este întotdeauna după vânzător
                new_filename = (
                    f"{vendor_clean}_"
                    f"{invoice_date}_"
                    f"{invoice_number_clean}.pdf"
                )

                own_company_normalized = normalize_company_name(
                    firma_proprie
                )

                vendor_normalized = normalize_company_name(
                    vendor
                )

                # Dacă firma proprie este vânzătorul,
                # factura merge la FacturiClienti.
                if (
                    own_company_normalized
                    and own_company_normalized in vendor_normalized
                ):
                    invoice_type = "CLIENT"
                    destination_folder = clienti_folder

                else:
                    invoice_type = "FURNIZOR"
                    destination_folder = furnizori_folder

                destination_path = get_unique_file_path(
                    destination_folder,
                    new_filename
                )

                shutil.copy2(
                    original_pdf_path,
                    destination_path
                )

                final_filename = os.path.basename(
                    destination_path
                )

                results.append({
                    "Fisier initial": uploaded_file.name,
                    "Fisier nou": final_filename,
                    "Tip": invoice_type,
                    "Vanzator": vendor,
                    "Cumparator": buyer,
                    "Data factura": invoice_date,
                    "Numar factura": invoice_number,
                    "Valoare neta": net_value,
                    "OCR valoare": value_ocr_text
                })

                logs.append(
                    f"✔ {uploaded_file.name} → "
                    f"{invoice_type} → {final_filename}"
                )

            except Exception as error:
                error_message = (
                    f"❌ Eroare la {uploaded_file.name}: {error}"
                )

                errors.append(error_message)
                logs.append(error_message)

            progress_bar.progress(
                index / number_of_files
            )

        status_text.empty()

        if results:
            excel_bytes = create_excel(results)
            zip_bytes = create_zip(processed_folder)

            st.session_state["invoice_results"] = results
            st.session_state["invoice_logs"] = logs
            st.session_state["invoice_errors"] = errors
            st.session_state["invoice_excel"] = excel_bytes
            st.session_state["invoice_zip"] = zip_bytes

            st.success(
                f"Procesarea a fost finalizată. "
                f"{len(results)} facturi au fost procesate."
            )

        else:
            st.session_state["invoice_results"] = []
            st.session_state["invoice_logs"] = logs
            st.session_state["invoice_errors"] = errors

            st.error(
                "Nicio factură nu a putut fi procesată."
            )


# ============================================================
# AFIȘAREA REZULTATELOR
# ============================================================

if st.session_state.get("invoice_results"):

    results = st.session_state["invoice_results"]

    st.divider()
    st.subheader("📊 Rezultatele procesării")

    display_data = []

    for result in results:
        display_data.append({
            "Fișier inițial": result["Fisier initial"],
            "Fișier nou": result["Fisier nou"],
            "Tip": result["Tip"],
            "Vânzător": result["Vanzator"],
            "Cumpărător": result["Cumparator"],
            "Data": result["Data factura"],
            "Număr factură": result["Numar factura"],
            "Valoare netă": result["Valoare neta"]
        })

    dataframe = pd.DataFrame(display_data)

    st.dataframe(
        dataframe,
        use_container_width=True,
        hide_index=True
    )

    total_net_value = sum(
        result["Valoare neta"]
        for result in results
    )

    client_invoices = sum(
        1
        for result in results
        if result["Tip"] == "CLIENT"
    )

    supplier_invoices = sum(
        1
        for result in results
        if result["Tip"] == "FURNIZOR"
    )

    column1, column2, column3, column4 = st.columns(4)

    column1.metric(
        "Facturi procesate",
        len(results)
    )

    column2.metric(
        "Facturi clienți",
        client_invoices
    )

    column3.metric(
        "Facturi furnizori",
        supplier_invoices
    )

    column4.metric(
        "Total valoare netă",
        f"{total_net_value:,.2f}"
    )

    st.subheader("⬇️ Descărcare rezultate")

    download_column1, download_column2 = st.columns(2)

    with download_column1:
        st.download_button(
            label="📊 Descarcă Excel",
            data=st.session_state["invoice_excel"],
            file_name="ValoriNet.xlsx",
            mime=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            use_container_width=True
        )

    with download_column2:
        st.download_button(
            label="🗂️ Descarcă facturile sortate",
            data=st.session_state["invoice_zip"],
            file_name="FacturiPrelucrate.zip",
            mime="application/zip",
            use_container_width=True
        )

    with st.expander("🔍 Vezi textul OCR pentru valorile nete"):
        for result in results:
            st.write(f"**{result['Fisier initial']}**")
            st.code(
                result["OCR valoare"]
                if result["OCR valoare"]
                else "Nu a fost detectat text."
            )


# ============================================================
# LOGURI ȘI ERORI
# ============================================================

if st.session_state.get("invoice_logs"):

    with st.expander("📋 Vezi logul procesării"):
        for log in st.session_state["invoice_logs"]:
            st.write(log)


if st.session_state.get("invoice_errors"):

    if st.session_state["invoice_errors"]:
        with st.expander("⚠️ Vezi erorile"):
            for error in st.session_state["invoice_errors"]:
                st.error(error)
